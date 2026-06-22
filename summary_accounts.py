# -*- coding: utf-8 -*-
import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

folder = r'C:\Users\没饭局\Desktop\1'
out_path = r'C:\Users\没饭局\Desktop\1\账号汇总表.xlsx'

# 只统计原始文件，排除合并后的文件
files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx') and not f.startswith('合并')])

type_names = {
    '天使': '天使黄金',
    '爆爽': '天使爆爽',
    '集结': '天使集结',
    '周年': '天使周年',
    '万魂': '天使万魂'
}

def get_type(name):
    for k, v in type_names.items():
        if k in name:
            return v
    return '其他'

def get_server_num(name):
    import re
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0

# 统计数据
summary = []
for fname in files:
    path = os.path.join(folder, fname)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = [r for r in rows[1:] if r[0] is not None]
    summary.append({
        'filename': fname,
        'type': get_type(fname),
        'num': get_server_num(fname),
        'count': len(data_rows)
    })

summary.sort(key=lambda x: (x['type'], x['num']))

total = sum(s['count'] for s in summary)

# 创建汇总表
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '汇总'

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
subtotal_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
total_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

headers = ['分类', '文件名', '账号数']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

row = 2
current_type = None
type_total = 0

for item in summary:
    t = item['type']
    if t != current_type:
        if current_type is not None:
            ws.cell(row=row, column=1, value='小计')
            ws.cell(row=row, column=3, value=type_total)
            for c in range(1, 4):
                ws.cell(row=row, column=c).border = border
                ws.cell(row=row, column=c).fill = subtotal_fill
                ws.cell(row=row, column=c).font = Font(bold=True)
            row += 1
            type_total = 0
        current_type = t
    type_total += item['count']
    ws.cell(row=row, column=1, value=t)
    ws.cell(row=row, column=2, value=item['filename'])
    ws.cell(row=row, column=3, value=item['count'])
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = border
    row += 1

# 最后一个小计
ws.cell(row=row, column=1, value='小计')
ws.cell(row=row, column=3, value=type_total)
for c in range(1, 4):
    ws.cell(row=row, column=c).border = border
    ws.cell(row=row, column=c).fill = subtotal_fill
    ws.cell(row=row, column=c).font = Font(bold=True)
row += 1

# 合计
ws.cell(row=row, column=1, value='合计')
ws.cell(row=row, column=3, value=total)
for c in range(1, 4):
    cell = ws.cell(row=row, column=c)
    cell.border = border
    cell.fill = total_fill
    cell.font = Font(bold=True, size=12)

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 25
ws.column_dimensions['C'].width = 12

wb.save(out_path)

print('=== 账号汇总表 ===')
print(f'{"分类":<10} {"文件名":<25} {"账号数":>8}')
print('-' * 48)
current_type = None
type_total = 0
for item in summary:
    t = item['type']
    if t != current_type:
        if current_type is not None:
            print(f'{"小计":<10} {current_type:<25} {type_total:>8}')
        current_type = t
        type_total = 0
    type_total += item['count']
    print(f'{item["type"]:<10} {item["filename"]:<25} {item["count"]:>8}')
print(f'{"小计":<10} {current_type:<25} {type_total:>8}')
print('-' * 48)
print(f'{"合计":<10} {"20个文件":<25} {total:>8}')
print(f'\n✅ 已保存: {out_path}')
