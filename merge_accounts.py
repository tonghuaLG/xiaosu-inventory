# -*- coding: utf-8 -*-
import openpyxl
import os

folder = r'C:\Users\没饭局\Desktop\1'
out_path = r'C:\Users\没饭局\Desktop\1\合并账号汇总.xlsx'
files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx')])

all_data = []
for fname in files:
    path = os.path.join(folder, fname)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    data_rows = [r for r in rows[1:] if r[0] is not None]
    print(f'{fname}: {len(data_rows)}')
    all_data.extend(data_rows)

print(f'\n总账号数: {len(all_data)}')

# 创建新的工作簿
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '合并账号'

# 写入表头
headers = ['账号', '密码', '小区', '状态', '版本']
ws_out.append(headers)

# 写入数据
for row in all_data:
    ws_out.append(list(row))

wb_out.save(out_path)
print(f'\n已保存到: {out_path}')
