# -*- coding: utf-8 -*-
import openpyxl
import os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

folder = r'C:\Users\没饭局\Desktop\1'
out_path = r'C:\Users\没饭局\Desktop\1\合并账号汇总.xlsx'

# 只统计原始文件
files = sorted([f for f in os.listdir(folder) if f.endswith('.xlsx') and not f.startswith('合并') and not f.startswith('账号')])

type_names = {
    '天使': '天使黄金',
    '爆爽': '天使爆爽',
    '集结': '天使集结',
    '周年': '天使周年',
    '万魂': '天使万魂'
}

def get_type(name):
    for k, v in type_names.items():
        if k in name:
            return v
    return '其他'

def get_server_num(name):
    import re
    m = re.search(r'(\d+)', name)
    return int(m.group(1)) if m else 0

# 读取所有数据
all_data = []
for fname in files:
    path = os.path.join(folder, fname)
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    for r in rows[1:]:
        if r[0] is not None:
            all_data.append({
                'type': get_type(fname),
                'num': get_server_num(fname),
                'row': r
            })

# 按类型+编号排序
all_data.sort(key=lambda x: (x['type'], x['num']))

# 创建新工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '合并账号'

# 样式
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
type_fills = {
    '天使黄金': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    '天使爆爽': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
    '天使集结': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
    '天使周年': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
    '天使万魂': PatternFill(start_color='EAD1DC', end_color='EAD1DC', fill_type='solid'),
}

# 表头
headers = ['账号', '密码', '小区', '状态', '版本']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 14

# 写入数据
row = 2
prev_type = None
for item in all_data:
    t = item['type']
    if t != prev_type:
        fill = type_fills.get(t)
        # 空一行分隔不同类型
        ws.cell(row=row, column=1, value=f'--- {t} ---')
        ws.cell(row=row, column=1).font = Font(bold=True, italic=True, color='808080')
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        row += 1
        prev_type = t

    fill = type_fills.get(t)
    ws.cell(row=row, column=1, value=item['row'][0])  # 账号
    ws.cell(row=row, column=2, value=item['row'][1])  # 密码
    ws.cell(row=row, column=3, value=item['row'][2] if len(item['row']) > 2 else None)  # 小区
    ws.cell(row=row, column=4, value=item['row'][3] if len(item['row']) > 3 else None)  # 状态
    ws.cell(row=row, column=5, value=item['row'][4] if len(item['row']) > 4 else None)  # 版本

    for c in range(1, 6):
        ws.cell(row=row, column=c).border = border
        if fill:
            ws.cell(row=row, column=c).fill = fill
    row += 1

wb.save(out_path)
print(f'Done! Total: {len(all_data)} accounts')
print(f'Saved to: {out_path}')
